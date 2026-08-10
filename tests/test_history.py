"""tests/test_history.py — 一時 SQLite DB を使った history モジュールの CRUD テスト"""

import io
import sys
from pathlib import Path

import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import history as hist
import routes.history as history_routes  # noqa: E402


@pytest.fixture(autouse=True)
def temp_db(tmp_path, monkeypatch):
    """各テストで新しい一時 DB を使用する"""
    db_file = tmp_path / "test_history.db"
    monkeypatch.setattr(hist, "DB_PATH", db_file)
    yield db_file


@pytest.fixture
def client():
    """/api/history/export 等のルートをテストするための TestClient"""
    app = FastAPI()
    app.include_router(history_routes.router)
    with TestClient(app) as c:
        yield c


# ------------------------------------------------------------------ #
# 基本 CRUD
# ------------------------------------------------------------------ #


def test_save_and_get():
    rowid = hist.save_history(
        positive="1girl, cute", negative="blurry", image_name="test.png", style="anime", tone="vibrant", quality="high"
    )
    assert isinstance(rowid, int)
    items = hist.get_history()
    assert len(items) == 1
    assert items[0]["positive"] == "1girl, cute"
    assert items[0]["image_name"] == "test.png"


def test_save_and_get_provider_model_and_template_metadata():
    rowid = hist.save_history(
        positive="a cat",
        negative="blurry",
        template="a {cat|dog}",
        provider="gemini",
        model="gemini-test",
    )

    item = hist.get_history_item(rowid)
    assert item["template"] == "a {cat|dog}"
    assert item["provider"] == "gemini"
    assert item["model"] == "gemini-test"


def test_get_with_limit():
    for i in range(5):
        hist.save_history(positive=f"prompt {i}", negative="neg")
    items = hist.get_history(limit=3)
    assert len(items) == 3


def test_get_with_offset():
    for i in range(5):
        hist.save_history(positive=f"item {i}", negative="neg")
    items_all = hist.get_history(limit=None)
    items_offset = hist.get_history(limit=None, offset=2)
    assert len(items_all) == 5
    assert len(items_offset) == 3


def test_limit_none_returns_all():
    for i in range(15):
        hist.save_history(positive=f"p{i}", negative="n")
    items = hist.get_history(limit=None)
    assert len(items) == 15


def test_search_filter():
    hist.save_history(positive="beautiful landscape", negative="ugly")
    hist.save_history(positive="anime girl", negative="bad")
    items = hist.get_history(search="landscape")
    assert len(items) == 1
    assert "landscape" in items[0]["positive"]


def test_search_with_percent_wildcard():
    """% を含む検索語はリテラルとして扱われ、全件マッチしない"""
    hist.save_history(positive="normal prompt", negative="bad")
    hist.save_history(positive="another prompt", negative="ugly")
    # "%" はエスケープされてリテラル % として検索される
    items = hist.get_history(search="%")
    assert len(items) == 0  # どちらの行も "%" を含まない


def test_search_with_underscore():
    """_ を含む検索語はリテラルとして扱われる"""
    hist.save_history(positive="has_underscore here", negative="bad")
    hist.save_history(positive="no special chars", negative="ugly")
    items = hist.get_history(search="_")
    assert len(items) == 1
    assert "has_underscore" in items[0]["positive"]


def test_style_filter():
    hist.save_history(positive="p1", negative="n", style="anime")
    hist.save_history(positive="p2", negative="n", style="photorealistic")
    items = hist.get_history(style="anime")
    assert len(items) == 1


def test_quality_filter():
    hist.save_history(positive="p1", negative="n", quality="high")
    hist.save_history(positive="p2", negative="n", quality="ultra")
    items = hist.get_history(quality="ultra")
    assert len(items) == 1


def test_favorites_only():
    rowid = hist.save_history(positive="fav", negative="n")
    hist.save_history(positive="not fav", negative="n")
    hist.toggle_favorite(rowid)
    items = hist.get_history(favorites_only=True)
    assert len(items) == 1
    assert items[0]["positive"] == "fav"


def test_toggle_favorite():
    rowid = hist.save_history(positive="test", negative="n")
    updated = hist.toggle_favorite(rowid)
    assert updated["is_favorite"] == 1
    updated2 = hist.toggle_favorite(rowid)
    assert updated2["is_favorite"] == 0


def test_toggle_favorite_not_found():
    result = hist.toggle_favorite(99999)
    assert result is None


def test_delete_history_item():
    rowid = hist.save_history(positive="to delete", negative="n")
    assert hist.delete_history_item(rowid) is True
    assert hist.get_history_item(rowid) is None


def test_delete_nonexistent():
    assert hist.delete_history_item(99999) is False


def test_clear_all_history():
    for i in range(3):
        hist.save_history(positive=f"p{i}", negative="n")
    count = hist.clear_all_history()
    assert count == 3
    assert hist.get_history() == []


def test_get_history_count():
    hist.save_history(positive="a", negative="n", style="anime")
    hist.save_history(positive="b", negative="n", style="photo")
    total = hist.get_history_count()
    assert total == 2
    filtered = hist.get_history_count(style="anime")
    assert filtered == 1


# ------------------------------------------------------------------ #
# Tags
# ------------------------------------------------------------------ #


def test_add_tags():
    rowid = hist.save_history(positive="tagged", negative="n")
    tags = hist.add_tags(rowid, ["portrait", "anime"])
    assert sorted(tags) == ["anime", "portrait"]


def test_add_tags_normalized():
    """Tags are lowercased and stripped"""
    rowid = hist.save_history(positive="tagged", negative="n")
    tags = hist.add_tags(rowid, ["  Portrait  ", "ANIME"])
    assert sorted(tags) == ["anime", "portrait"]


def test_add_duplicate_tags():
    """Adding duplicate tags is idempotent"""
    rowid = hist.save_history(positive="tagged", negative="n")
    hist.add_tags(rowid, ["portrait"])
    tags = hist.add_tags(rowid, ["portrait", "anime"])
    assert sorted(tags) == ["anime", "portrait"]


def test_remove_tag():
    rowid = hist.save_history(positive="tagged", negative="n")
    hist.add_tags(rowid, ["portrait", "anime"])
    remaining = hist.remove_tag(rowid, "portrait")
    assert remaining == ["anime"]


def test_get_tags_empty():
    rowid = hist.save_history(positive="no tags", negative="n")
    assert hist.get_tags(rowid) == []


def test_get_all_tags():
    r1 = hist.save_history(positive="p1", negative="n")
    r2 = hist.save_history(positive="p2", negative="n")
    hist.add_tags(r1, ["portrait", "anime"])
    hist.add_tags(r2, ["portrait"])
    all_tags = hist.get_all_tags()
    assert len(all_tags) == 2
    assert all_tags[0]["tag"] == "portrait"
    assert all_tags[0]["count"] == 2
    assert all_tags[1]["tag"] == "anime"
    assert all_tags[1]["count"] == 1


def test_history_filter_by_tag():
    r1 = hist.save_history(positive="p1", negative="n")
    hist.save_history(positive="p2", negative="n")
    hist.add_tags(r1, ["portrait"])
    items = hist.get_history(tag="portrait")
    assert len(items) == 1
    assert items[0]["positive"] == "p1"


def test_history_count_with_tag():
    r1 = hist.save_history(positive="p1", negative="n")
    hist.save_history(positive="p2", negative="n")
    hist.add_tags(r1, ["portrait"])
    assert hist.get_history_count(tag="portrait") == 1
    assert hist.get_history_count() == 2


def test_history_items_include_tags():
    """get_history returns items with 'tags' key"""
    rowid = hist.save_history(positive="with tags", negative="n")
    hist.add_tags(rowid, ["landscape", "hdr"])
    items = hist.get_history()
    assert "tags" in items[0]
    assert sorted(items[0]["tags"]) == ["hdr", "landscape"]


def test_delete_history_cascades_tags():
    """Deleting a history item should cascade-delete its tags"""
    rowid = hist.save_history(positive="to delete", negative="n")
    hist.add_tags(rowid, ["portrait"])
    hist.delete_history_item(rowid)
    assert hist.get_tags(rowid) == []


# ------------------------------------------------------------------ #
# Export — XLSX
# ------------------------------------------------------------------ #

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_export_xlsx_returns_200_with_correct_headers(client):
    hist.save_history(positive="1girl, cute", negative="blurry")
    response = client.get("/api/history/export", params={"format": "xlsx"})
    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX_MEDIA_TYPE
    assert response.headers["content-disposition"] == 'attachment; filename="prompt_history.xlsx"'


def test_export_xlsx_workbook_matches_history(client):
    r1 = hist.save_history(positive="1girl, cute", negative="blurry", image_name="a.png")
    hist.save_history(positive="landscape", negative="ugly", image_name="b.png")
    hist.add_tags(r1, ["portrait", "anime"])

    items = hist.get_history(limit=None)
    response = client.get("/api/history/export", params={"format": "xlsx"})
    assert response.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb["History"]
    rows = list(ws.iter_rows(values_only=True))

    expected_header = list(items[0].keys())
    assert list(rows[0]) == expected_header
    assert len(rows) - 1 == len(items)  # header + one row per item

    tags_col = expected_header.index("tags")
    row_by_id = {row[expected_header.index("id")]: row for row in rows[1:]}
    assert row_by_id[r1][tags_col] == "anime, portrait"


def test_export_xlsx_empty_history_returns_valid_workbook(client):
    response = client.get("/api/history/export", params={"format": "xlsx"})
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb["History"]
    assert ws.max_row == 1
    assert ws["A1"].value is None


# ------------------------------------------------------------------ #
# GET /api/history — limit/offset validation
# ------------------------------------------------------------------ #


def test_get_history_negative_limit_returns_422(client):
    response = client.get("/api/history", params={"limit": -1})
    assert response.status_code == 422


def test_get_history_zero_limit_returns_422(client):
    response = client.get("/api/history", params={"limit": 0})
    assert response.status_code == 422


def test_get_history_too_large_limit_returns_422(client):
    response = client.get("/api/history", params={"limit": 100000})
    assert response.status_code == 422


def test_get_history_negative_offset_returns_422(client):
    response = client.get("/api/history", params={"offset": -1})
    assert response.status_code == 422


def test_get_history_valid_limit_and_offset_paginates(client):
    for i in range(5):
        hist.save_history(positive=f"item {i}", negative="n")
    response = client.get("/api/history", params={"limit": 2, "offset": 0})
    assert response.status_code == 200
    data = response.json()
    assert len(data["items"]) == 2
    assert data["total"] == 5

    response2 = client.get("/api/history", params={"limit": 2, "offset": 2})
    assert response2.status_code == 200
    data2 = response2.json()
    assert len(data2["items"]) == 2
    # Offset pages should not overlap
    ids_page1 = {item["id"] for item in data["items"]}
    ids_page2 = {item["id"] for item in data2["items"]}
    assert ids_page1.isdisjoint(ids_page2)


def test_get_history_max_allowed_limit_is_accepted(client):
    response = client.get("/api/history", params={"limit": 200})
    assert response.status_code == 200


# ------------------------------------------------------------------ #
# Batched tag lookup (guards against the N+1 query refactor)
# ------------------------------------------------------------------ #


def test_get_history_batch_tags_for_many_items():
    """Tags are attached correctly for a batch of items, including items
    with no tags and items with multiple tags."""
    r1 = hist.save_history(positive="no tags here", negative="n")
    r2 = hist.save_history(positive="one tag", negative="n")
    r3 = hist.save_history(positive="many tags", negative="n")
    hist.add_tags(r2, ["portrait"])
    hist.add_tags(r3, ["zebra", "anime", "hdr"])

    items = hist.get_history(limit=None)
    by_id = {item["id"]: item for item in items}

    assert by_id[r1]["tags"] == []
    assert by_id[r2]["tags"] == ["portrait"]
    # tags within an item are alphabetically ordered
    assert by_id[r3]["tags"] == ["anime", "hdr", "zebra"]


def test_get_history_batch_tags_chunking_beyond_500_ids():
    """The tag lookup chunks IDs; verify it still returns correct tags when
    the number of history rows exceeds the chunk size."""
    ids = []
    for i in range(520):
        ids.append(hist.save_history(positive=f"p{i}", negative="n"))
    hist.add_tags(ids[0], ["first"])
    hist.add_tags(ids[-1], ["last"])

    items = hist.get_history(limit=None)
    by_id = {item["id"]: item for item in items}
    assert by_id[ids[0]]["tags"] == ["first"]
    assert by_id[ids[-1]]["tags"] == ["last"]
    # Everything else has no tags
    untagged = [i for i in ids[1:-1]]
    assert all(by_id[i]["tags"] == [] for i in untagged)


# ------------------------------------------------------------------ #
# init_db() caching (Task 3) does not leak between monkeypatched paths
# ------------------------------------------------------------------ #


def test_init_db_works_against_freshly_monkeypatched_db_path(tmp_path, monkeypatch):
    """Even though init_db() short-circuits after the first successful run for
    a given DB_PATH, pointing DB_PATH at a brand new path must still get a
    full initialization (tables created) rather than silently no-op-ing."""
    other_db = tmp_path / "another_history.db"
    monkeypatch.setattr(hist, "DB_PATH", other_db)

    # DB file doesn't exist yet; init_db() must create schema so this works.
    rowid = hist.save_history(positive="fresh db", negative="n")
    items = hist.get_history()
    assert len(items) == 1
    assert items[0]["id"] == rowid

    # Calling init_db() again for the same (already-initialized) path is a
    # cheap no-op and must not raise or break anything.
    hist.init_db()
    assert hist.get_history() == items
